import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from database.models import User, Subscriptions, TestAnswers

# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

def adjust_column_width(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try: 
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

def apply_header_style(sheet):
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

def apply_body_style(sheet):
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            # Agar matn juda uzun bo'lsa wrap text qilish mumkin, 
            # lekin hozircha oddiy markazlashtirish kifoya.

async def generate_subscribers_report():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Obunachilar"
    
    headers = ["ID", "User ID", "Ism-familiya", "Sana"]
    sheet.append(headers)
    
    subscriptions = await Subscriptions.all()
    # Foydalanuvchilarni olish va lug'at shakliga keltirish (tezroq qidirish uchun)
    users = await User.filter(tg_id__in=[sub.user_id for sub in subscriptions])
    users_dict = {user.tg_id: user.full_name for user in users}
    
    for sub in subscriptions:
        full_name = users_dict.get(sub.user_id, "Noma'lum")
        sheet.append([sub.id, sub.user_id, full_name, sub.created_at.strftime("%Y-%m-%d %H:%M:%S")])
        
    apply_header_style(sheet)
    apply_body_style(sheet)
    adjust_column_width(sheet)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer

async def generate_members_report():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "A'zolar"
    
    headers = ["ID", "Telegram ID", "Ism-familiya"]
    sheet.append(headers)
    
    users = await User.all()
    
    for user in users:
        sheet.append([user.id, user.tg_id, user.full_name])
        
    apply_header_style(sheet)
    apply_body_style(sheet)
    adjust_column_width(sheet)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer

async def generate_rating_report():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Reyting"
    
    headers = ["O'rin", "Ism-familiya", "Jami ball", "Yechilgan testlar"]
    sheet.append(headers)
    
    users = await User.all()
    user_stats = []
    
    for user in users:
        answers = await TestAnswers.filter(user_id=user.tg_id)
        total_score = sum(a.score for a in answers)
        tests_count = len(answers)
        if tests_count > 0:
            user_stats.append({
                "name": user.full_name,
                "score": total_score,
                "count": tests_count
            })
    
    # Sort by score descending
    user_stats.sort(key=lambda x: x["score"], reverse=True)
    
    for index, stat in enumerate(user_stats, 1):
        sheet.append([index, stat["name"], stat["score"], stat["count"]])
        
    apply_header_style(sheet)
    apply_body_style(sheet)
    adjust_column_width(sheet)
    
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
